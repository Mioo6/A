"""
excel_exporter.py
=================
Exporte un DataFrame classifié vers un fichier .xlsx multi-feuilles,
avec une mise en forme reproduisant exactement Tri_Elements_Revit.xlsx :

  Feuille 1  📊 Synthèse           — tableau récapitulatif + descriptions
  Feuille 2  ♻ Réemploi            — palette verte
  Feuille 3  🔄 Recyclage          — palette bleue
  Feuille 4  🗑 Autres déchets     — palette rouge

L'ordre des colonnes par feuille est :
  Famille | Famille et type | Type | Matériau | Niveau | ID Type | ID
"""
from __future__ import annotations

import io
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from .admin_data import PALETTES
from .classifier import summarize_classification


# Colonnes de sortie standard (ordre fixe, comme la référence)
OUTPUT_COLUMNS = ["Famille", "Famille et type", "Type", "Matériau",
                  "Niveau", "ID Type", "ID"]

# Largeurs de colonnes (issues de l'analyse du fichier de référence)
COLUMN_WIDTHS = {
    "Famille": 28.0,
    "Famille et type": 50.0,
    "Type": 45.0,
    "Matériau": 30.0,
    "Niveau": 28.0,
    "ID Type": 12.0,
    "ID": 10.0,
}

# Bordure standard (trait fin gris foncé)
THIN_BORDER = Border(
    top=Side(style="thin", color="B0BEC5"),
    bottom=Side(style="thin", color="B0BEC5"),
    left=Side(style="thin", color="B0BEC5"),
    right=Side(style="thin", color="B0BEC5"),
)


# =============================================================================
# Préparation du DataFrame de sortie
# =============================================================================

def _prepare_output_df(df: pd.DataFrame, categorie: str) -> pd.DataFrame:
    """Filtre et reformate les colonnes pour la feuille d'une catégorie."""
    sub = df.loc[df["Categorie"] == categorie].copy()
    # Construction des colonnes attendues (même si certaines sont absentes)
    out = pd.DataFrame()
    out["Famille"] = sub.get("Famille", "")
    out["Famille et type"] = sub.get("Famille et type", "")
    out["Type"] = sub.get("Type", "")
    out["Matériau"] = ""  # non extrait directement de Revit, laissé vide
    out["Niveau"] = sub.get("Niveau", "")
    out["ID Type"] = ""  # non disponible dans les CSV Revit standard
    out["ID"] = sub.get("Identifiant", "")
    return out.reset_index(drop=True)


# =============================================================================
# Construction des feuilles
# =============================================================================

def _write_synthesis_sheet(ws, summary, title_main: str):
    """Écrit la feuille 📊 Synthèse."""
    pal = PALETTES["Synthese"]

    # En-tête principal (ligne 1)
    ws.merge_cells("A1:C1")
    ws["A1"] = title_main
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=pal["primary"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # Sous-titre (ligne 2)
    ws.merge_cells("A2:C2")
    ws["A2"] = "Classification des éléments selon leur fin de vie"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="546E7A")
    ws["A2"].fill = PatternFill("solid", fgColor=pal["soft"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Ligne vide
    ws.row_dimensions[3].height = 10

    # En-tête tableau (ligne 4)
    headers = ["Catégorie", "Nombre d'éléments", "Proportion"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=pal["secondary"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[4].height = 24

    # Lignes de données (5-7)
    rows_data = [
        ("Reemploi",      "♻  Réemploi",      summary.n_reemploi,  summary.pct_reemploi),
        ("Recyclage",     "🔄  Recyclage",     summary.n_recyclage, summary.pct_recyclage),
        ("AutresDechets", "🗑  Autres déchets", summary.n_dechets,   summary.pct_dechets),
    ]
    for i, (cat_key, label, count, pct) in enumerate(rows_data):
        r = 5 + i
        cat_pal = PALETTES[cat_key]
        # Colonne A : libellé
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Arial", size=11, bold=True, color=cat_pal["primary"])
        c.fill = PatternFill("solid", fgColor=cat_pal["soft"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        # Colonne B : nombre
        c = ws.cell(row=r, column=2, value=count)
        c.font = Font(name="Arial", size=11, bold=True, color=cat_pal["primary"])
        c.fill = PatternFill("solid", fgColor=cat_pal["soft"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        # Colonne C : pourcentage
        c = ws.cell(row=r, column=3, value=f"{pct:.1f}%")
        c.font = Font(name="Arial", size=11, bold=True, color=cat_pal["primary"])
        c.fill = PatternFill("solid", fgColor=cat_pal["soft"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        ws.row_dimensions[r].height = 28

    # Ligne TOTAL (8)
    c = ws.cell(row=8, column=1, value="TOTAL")
    c.font = Font(name="Arial", size=12, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=pal["primary"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    c = ws.cell(row=8, column=2, value=summary.total)
    c.font = Font(name="Arial", size=12, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=pal["primary"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    c = ws.cell(row=8, column=3, value="100%")
    c.font = Font(name="Arial", size=12, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=pal["primary"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    ws.row_dimensions[8].height = 28

    # Lignes vides
    ws.row_dimensions[9].height = 14
    ws.row_dimensions[10].height = 10

    # Descriptions des catégories (lignes 11-13)
    descriptions = [
        ("Reemploi",
         "♻  RÉEMPLOI",
         "Composants intègres réutilisables tels quels : menuiseries, "
         "panneaux système, mur-rideau, baies libres, garde-corps, "
         "structures bois lamellé-collé, escaliers préfabriqués."),
        ("Recyclage",
         "🔄  RECYCLAGE",
         "Béton, brique, métal, bois de structure, granulats — valorisables "
         "via filières de recyclage matière (Valobat, Paprec, Ecomaison)."),
        ("AutresDechets",
         "🗑  AUTRES DÉCHETS",
         "Isolants dégradés, cloisons plâtre, composites non séparables — "
         "à orienter vers déchèterie ou incinération (Tri'n'Collect)."),
    ]
    for i, (cat_key, label, desc) in enumerate(descriptions):
        r = 11 + i
        cat_pal = PALETTES[cat_key]
        # Colonne A : titre catégorie
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=cat_pal["primary"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Colonnes B+C fusionnées : description
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=desc)
        c.font = Font(name="Arial", size=9, color="37474F")
        c.fill = PatternFill("solid", fgColor=cat_pal["soft"])
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 36

    # Largeurs de colonnes
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18

    # Pas d'affichage des grilles
    ws.sheet_view.showGridLines = False


def _write_category_sheet(ws, df_cat: pd.DataFrame, categorie: str):
    """Écrit une feuille de catégorie (Réemploi / Recyclage / AutresDechets)."""
    pal = PALETTES[categorie]
    n_rows = len(df_cat)

    # En-tête principal (ligne 1)
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{pal['icon']}  {pal['label']} — {pal['subtitle']}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=pal["primary"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sous-titre avec total (ligne 2)
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Total : {n_rows} élément(s)"
    ws["A2"].font = Font(name="Arial", size=11, italic=True, color=pal["primary"])
    ws["A2"].fill = PatternFill("solid", fgColor=pal["soft"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # En-têtes de colonnes (ligne 3)
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        c = ws.cell(row=3, column=col_idx, value=col_name)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=pal["secondary"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[3].height = 22

    # Données (à partir de ligne 4) avec alternance de couleurs
    for row_idx, (_, row) in enumerate(df_cat.iterrows()):
        excel_row = 4 + row_idx
        is_alt = row_idx % 2 == 0  # lignes paires colorées
        bg_color = pal["soft"] if is_alt else "FFFFFF"
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = row[col_name] if col_name in df_cat.columns else ""
            c = ws.cell(row=excel_row, column=col_idx, value=value)
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg_color)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = THIN_BORDER
        ws.row_dimensions[excel_row].height = 16

    # Largeurs de colonnes
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_name]

    # Freeze panes : on fige sous l'en-tête (ligne 4 visible toujours en haut)
    ws.freeze_panes = "A4"

    # Pas d'affichage des grilles
    ws.sheet_view.showGridLines = False


# =============================================================================
# API publique
# =============================================================================

def build_xlsx(df_classified: pd.DataFrame, title_main: str) -> bytes:
    """Construit le fichier Excel complet et renvoie ses bytes.

    Args:
        df_classified: DataFrame déjà classifié (colonne 'Categorie').
        title_main: titre à afficher en haut de la feuille de synthèse.

    Returns:
        Contenu binaire du fichier .xlsx, prêt à être servi via
        st.download_button() ou écrit sur disque.
    """
    summary = summarize_classification(df_classified)

    wb = Workbook()
    # On supprime la feuille par défaut
    wb.remove(wb.active)

    # Feuille 1 : 📊 Synthèse
    ws_synth = wb.create_sheet("📊 Synthèse")
    _write_synthesis_sheet(ws_synth, summary, title_main)

    # Feuille 2 : ♻ Réemploi
    df_re = _prepare_output_df(df_classified, "Reemploi")
    ws_re = wb.create_sheet("♻ Réemploi")
    _write_category_sheet(ws_re, df_re, "Reemploi")

    # Feuille 3 : 🔄 Recyclage
    df_rc = _prepare_output_df(df_classified, "Recyclage")
    ws_rc = wb.create_sheet("🔄 Recyclage")
    _write_category_sheet(ws_rc, df_rc, "Recyclage")

    # Feuille 4 : 🗑 Autres déchets
    df_de = _prepare_output_df(df_classified, "AutresDechets")
    ws_de = wb.create_sheet("🗑 Autres déchets")
    _write_category_sheet(ws_de, df_de, "AutresDechets")

    # Feuille active à l'ouverture = Synthèse
    wb.active = 0

    # Sérialisation en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
